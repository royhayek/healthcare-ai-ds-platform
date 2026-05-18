"""Predictions artifact - Excel with conditional formatting (§4.5).

Generates an Excel workbook containing:
- All original input columns
- prediction, probability, confidence_band, similarity_score
- top_3_drivers / top_3_dampeners (from per-row SHAP)
- decision_at_optimal_threshold
- risk_flag

Uses the test-set data (X_test + y_test from splits.pkl) since full inference
predictions are handled by the prediction endpoint in Step 8.
"""

from __future__ import annotations

import io
import logging
import pickle
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd

from backend.deliverables.base import CLINICAL_DISCLAIMER_SHORT, GeneratedDeliverable

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession
    from backend.core.database import Dataset, Run

logger = logging.getLogger(__name__)

_CONFIDENCE_HIGH = 0.75
_CONFIDENCE_LOW = 0.40

# Clinical risk tier thresholds (screening / FN-heavy context)
_RISK_LOW_MAX = 0.15
_RISK_MEDIUM_MAX = 0.40


def _confidence_band(score: float) -> str:
    if score >= _CONFIDENCE_HIGH:
        return "high"
    if score >= _CONFIDENCE_LOW:
        return "medium"
    return "low"


def _clinical_risk_tier(probability: float) -> str:
    """Map calibrated probability to clinical risk tier label."""
    if probability < _RISK_LOW_MAX:
        return "Low Risk"
    if probability < _RISK_MEDIUM_MAX:
        return "Medium Risk"
    return "High Risk"


def _top_drivers(shap_row: np.ndarray, feature_names: list[str], k: int = 3) -> tuple[str, str]:
    """Return top-k positive and negative SHAP contributors for one row."""
    if len(shap_row) == 0 or not feature_names:
        return "", ""
    pairs = list(zip(feature_names, shap_row.tolist()))
    pairs.sort(key=lambda x: x[1], reverse=True)
    drivers = [f for f, v in pairs[:k] if v > 0]
    dampeners = [f for f, v in pairs[-k:][::-1] if v < 0]
    return ", ".join(drivers), ", ".join(dampeners)


async def generate_predictions_artifact(
    run: "Run",
    dataset: "Dataset | None",
    session: "AsyncSession",
    ctx: dict[str, Any],
) -> GeneratedDeliverable:
    """Generate the predictions Excel artifact."""
    from backend.core.storage import storage

    # Load model and test splits
    model_path = run.model_storage_path
    splits_path = f"runs/{run.id}/splits.pkl"

    try:
        model_bytes = await storage.download(model_path)
        import sys
        import joblib
        _prev_limit = sys.getrecursionlimit()
        sys.setrecursionlimit(10000)
        try:
            model = joblib.load(io.BytesIO(model_bytes))
        finally:
            sys.setrecursionlimit(_prev_limit)
    except Exception as exc:
        logger.warning("Could not load model: %s", exc)
        # Return empty artifact rather than crashing the whole bundle
        return _empty_artifact(run.id)

    try:
        splits_bytes = await storage.download(splits_path)
        splits = pickle.loads(splits_bytes)  # nosec - internal artifact
        X_test: pd.DataFrame = splits["X_test"]
        y_test: pd.Series = splits["y_test"]
    except Exception as exc:
        logger.warning("Could not load splits: %s", exc)
        return _empty_artifact(run.id)

    task_type = ctx.get("task_type", "binary_classification")
    optimal_threshold = (
        ctx.get("threshold", {}).get("optimal") or 0.5
    )
    threshold_config = run.threshold_config or {}
    threshold_used = threshold_config.get("optimal_threshold", optimal_threshold)

    # Compute predictions - suppress sklearn feature-name warnings for models
    # trained before the set_output("default") fix was applied to cleaner.py.
    import warnings
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*does not have valid feature names.*",
                category=UserWarning,
            )
            if task_type in ("binary_classification", "multiclass"):
                y_proba = model.predict_proba(X_test)
                if task_type == "binary_classification":
                    proba_1d = y_proba[:, 1]
                    y_pred = (proba_1d >= threshold_used).astype(int)
                    probabilities = proba_1d.tolist()
                else:
                    y_pred = model.predict(X_test)
                    probabilities = y_proba.max(axis=1).tolist()
            else:
                y_pred = model.predict(X_test)
                probabilities = y_pred.tolist()
    except Exception as exc:
        logger.warning("Prediction failed: %s", exc)
        return _empty_artifact(run.id)

    # Per-row SHAP (limited to MAX_SAMPLES)
    MAX_SAMPLES = 500
    shap_drivers_col: list[str] = []
    shap_dampeners_col: list[str] = []
    per_row_shap: np.ndarray | None = None

    if len(X_test) <= MAX_SAMPLES:
        try:
            from backend.ml.explainer import compute_shap
            shap_result = compute_shap(
                model, X_test, list(X_test.columns), task_type,
                background_data=None,
            )
            if hasattr(shap_result, "shap_values_matrix") and shap_result.shap_values_matrix is not None:
                per_row_shap = shap_result.shap_values_matrix
        except Exception as exc:
            logger.debug("Per-row SHAP failed (non-fatal): %s", exc)

    feature_names = list(X_test.columns)
    n_rows = len(X_test)

    for i in range(n_rows):
        if per_row_shap is not None and i < len(per_row_shap):
            drivers, dampeners = _top_drivers(per_row_shap[i], feature_names)
        else:
            drivers, dampeners = "", ""
        shap_drivers_col.append(drivers)
        shap_dampeners_col.append(dampeners)

    # Similarity scores
    sim_scores: list[float] = []
    if run.faiss_index_path:
        try:
            sim_bytes = await storage.download(run.faiss_index_path)
            from backend.ml.similarity import SimilarityIndex
            sim_idx = SimilarityIndex.deserialize(sim_bytes)
            from backend.ml.cleaner import apply_preprocessor, build_preprocessor
            from backend.models.strategy import PreprocessingStrategy
            prep_strategy = PreprocessingStrategy.model_validate(run.preprocessing_strategy)
            sim_prep = build_preprocessor(prep_strategy, X_test)
            X_test_t, _ = apply_preprocessor(sim_prep, X_test, X_test)
            scores = sim_idx.score(X_test_t.to_numpy(dtype=float))
            sim_scores = scores.tolist()
        except Exception as exc:
            logger.debug("Similarity scoring failed (non-fatal): %s", exc)

    if not sim_scores:
        sim_scores = [1.0] * n_rows

    # Build result DataFrame
    proba_list = [float(p) for p in probabilities]
    df_out = X_test.copy().reset_index(drop=True)
    df_out["risk_score_pct"] = [f"{p * 100:.1f}%" for p in proba_list]
    df_out["clinical_risk_tier"] = [_clinical_risk_tier(p) for p in proba_list]
    df_out["prediction"] = np.asarray(y_pred)
    df_out["probability"] = proba_list
    df_out["similarity_score"] = sim_scores
    df_out["model_confidence"] = [_confidence_band(s) for s in sim_scores]
    df_out["top_3_risk_factors"] = shap_drivers_col
    df_out["top_3_protective_factors"] = shap_dampeners_col
    df_out["clinical_threshold_used"] = threshold_used
    df_out["actual_outcome"] = y_test.reset_index(drop=True)
    df_out["out_of_distribution_flag"] = (df_out["similarity_score"] < _CONFIDENCE_LOW).astype(int)
    df_out["clinical_disclaimer"] = CLINICAL_DISCLAIMER_SHORT

    # Write Excel, CSV, and Parquet (§4.5)
    excel_bytes = _write_excel(df_out, run.id)
    csv_bytes = df_out.to_csv(index=False).encode("utf-8")
    parquet_bytes = _write_parquet(df_out)

    # Store CSV and Parquet alongside the main artifact
    csv_path = f"runs/{run.id}/predictions.csv"
    parquet_path = f"runs/{run.id}/predictions.parquet"
    await storage.upload(csv_path, csv_bytes, "text/csv")
    await storage.upload(parquet_path, parquet_bytes, "application/octet-stream")

    return GeneratedDeliverable.build(
        name="predictions",
        fmt="xlsx",
        content=excel_bytes,
        run_id=run.id,
        inputs_used=["model", "test_splits", "shap_summary", "similarity_index"],
        audience="clinician, clinical analyst, care team",
    )


def _write_excel(df: pd.DataFrame, run_id: str) -> bytes:
    buf = io.BytesIO()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=9)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(bottom=thin)

        cols = list(df.columns)
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(val, float) and not pd.isna(val):
                    cell.value = round(val, 6)
                elif pd.isna(val) if isinstance(val, float) else False:
                    cell.value = ""
                else:
                    cell.value = val

        from openpyxl.formatting.rule import CellIsRule
        n_data_rows = len(df) + 1

        # Conditional formatting on clinical_risk_tier column
        try:
            tier_col_idx = cols.index("clinical_risk_tier") + 1
            tier_col_letter = get_column_letter(tier_col_idx)
            ws.conditional_formatting.add(
                f"{tier_col_letter}2:{tier_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"Low Risk"'],
                           fill=PatternFill(bgColor="D1FAE5")),  # green
            )
            ws.conditional_formatting.add(
                f"{tier_col_letter}2:{tier_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"Medium Risk"'],
                           fill=PatternFill(bgColor="FEF3C7")),  # amber
            )
            ws.conditional_formatting.add(
                f"{tier_col_letter}2:{tier_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"High Risk"'],
                           fill=PatternFill(bgColor="FEE2E2")),  # red
            )
        except Exception:
            pass

        # Conditional formatting on model_confidence column
        try:
            cb_col_idx = cols.index("model_confidence") + 1
            cb_col_letter = get_column_letter(cb_col_idx)
            ws.conditional_formatting.add(
                f"{cb_col_letter}2:{cb_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"high"'],
                           fill=PatternFill(bgColor="D1FAE5")),
            )
            ws.conditional_formatting.add(
                f"{cb_col_letter}2:{cb_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"medium"'],
                           fill=PatternFill(bgColor="FEF3C7")),
            )
            ws.conditional_formatting.add(
                f"{cb_col_letter}2:{cb_col_letter}{n_data_rows}",
                CellIsRule(operator="equal", formula=['"low"'],
                           fill=PatternFill(bgColor="FEE2E2")),
            )
        except Exception:
            pass

        # Auto-fit columns (approximate)
        for col_idx, col_name in enumerate(cols, start=1):
            max_len = max(len(str(col_name)), 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

        wb.save(buf)
    except Exception as exc:
        logger.warning("Excel formatting failed, writing plain CSV fallback: %s", exc)
        buf = io.BytesIO(df.to_csv(index=False).encode())

    return buf.getvalue()


def _write_parquet(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    if _has_pyarrow():
        df.to_parquet(buf, index=False, engine="pyarrow")
    else:
        try:
            df.to_parquet(buf, index=False, engine="fastparquet")
        except Exception:
            # Neither parquet engine available - fall back to gzip-compressed CSV
            buf = io.BytesIO(df.to_csv(index=False).encode("utf-8"))
    return buf.getvalue()


def _has_pyarrow() -> bool:
    try:
        import pyarrow  # noqa: F401
        return True
    except ImportError:
        return False


def _empty_artifact(run_id: str) -> GeneratedDeliverable:
    placeholder = b"run_id,error\n" + run_id.encode() + b",model or splits not available\n"
    return GeneratedDeliverable.build(
        name="predictions",
        fmt="xlsx",
        content=placeholder,
        run_id=run_id,
        inputs_used=[],
    )
